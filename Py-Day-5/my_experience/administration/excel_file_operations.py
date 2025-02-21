import openpyxl

def write_excel(filename, data):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for row in data:
            sheet.append(row)

        workbook.save(filename)
    except:
        return "Error writing to the Excel file!"

def read_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        data = [row for row in sheet.iter_rows(values_only=True)]
        return data
    except:
        return "Error reading the Excel file!"
